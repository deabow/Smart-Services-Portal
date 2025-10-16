from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand

from achievements.models import Achievement


class Command(BaseCommand):
    help = (
        "Export achievements to an Excel file with order, project, area, village, "
        "plus image links, and an Images sheet. Optionally embed thumbnails."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default=str(Path.cwd() / "engaz.xlsx"),
            help="Path to write the Excel file",
        )
        parser.add_argument(
            "--embed-images",
            action="store_true",
            help="Embed small thumbnails of the first image into the main sheet",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
        except ImportError as exc:
            raise SystemExit(
                "openpyxl is required. Please install it or run: pip install openpyxl"
            ) from exc

        output_path = Path(options["output"]).resolve()
        embed_images: bool = bool(options.get("embed_images"))

        achievements = (
            Achievement.objects.all()
            .order_by("id")
            .prefetch_related("images")
        )

        wb = Workbook()

        # Main sheet
        ws = wb.active
        ws.title = "Achievements"

        # Header row (Arabic)
        ws.append([
            "م",  # order
            "اسم المشروع",  # project name (title)
            "الوصف",
            "المركز",
            "القرية",
            "رابط الصورة الأولى",  # hyperlink to first image file (if exists)
        ])

        # Optional column for embedded thumbnails will be the next column
        thumb_col_idx = 7  # Column G if embedding

        for idx, achievement in enumerate(achievements, start=1):
            order_num = achievement.id  # stable current order by id
            project_name = achievement.title
            description = achievement.description
            center = achievement.area
            village = achievement.village or ""

            # Determine first image and hyperlink
            first_image_obj = None
            first_image_path = None
            if hasattr(achievement, "images"):
                first_image_obj = next(iter(achievement.images.all()), None)
                if first_image_obj and getattr(first_image_obj.image, "path", None):
                    first_image_path = Path(first_image_obj.image.path)

            row_values = [order_num, project_name, description, center, village, ""]
            ws.append(row_values)

            # Add hyperlink to first image if present
            if first_image_path and first_image_path.exists():
                # Set a file hyperlink (works in Excel on Windows)
                cell = ws.cell(row=ws.max_row, column=6)
                cell.value = "فتح الصورة"
                cell.hyperlink = first_image_path.as_uri() if first_image_path.drive else f"file:///{first_image_path}"
                cell.style = "Hyperlink"

            # Embed thumbnail if requested and image exists
            if embed_images and first_image_path and first_image_path.exists():
                try:
                    img = XLImage(str(first_image_path))
                    # Resize roughly to thumbnail
                    img.width = 96
                    img.height = 96
                    anchor = f"{self._col_letter(thumb_col_idx)}{ws.max_row}"
                    ws.add_image(img, anchor)
                    # Increase row height for thumbnail visibility
                    ws.row_dimensions[ws.max_row].height = 80
                except Exception:
                    # If any error occurs loading the image, skip embedding
                    pass

        # Images sheet with one row per image
        ws_images = wb.create_sheet(title="Images")
        ws_images.append([
            "م",  # order (achievement id)
            "اسم المشروع",
            "اسم الملف",
            "المسار الكامل",
            "رابط الملف",
        ])

        for achievement in achievements:
            project_name = achievement.title
            for img in achievement.images.all():
                img_path = getattr(img.image, "path", "") or ""
                link = ""
                if img_path:
                    p = Path(img_path)
                    if p.exists():
                        link = p.as_uri() if p.drive else f"file:///{p}"
                ws_images.append([
                    achievement.id,
                    project_name,
                    Path(img.image.name).name if getattr(img, "image", None) else "",
                    img_path,
                    link,
                ])

        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"Excel written to: {output_path}"))

    def _col_letter(self, idx: int) -> str:
        # Convert 1-based index to Excel column letters
        letters = []
        while idx > 0:
            idx, rem = divmod(idx - 1, 26)
            letters.append(chr(65 + rem))
        return "".join(reversed(letters))


